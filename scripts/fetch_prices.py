#!/usr/bin/env python3
"""Fetch daily fuel prices from ena.lt and produce JSON files."""

import json
import re
import shutil
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
HISTORY_DIR = DATA_DIR / "history"
DOCS_DATA = ROOT / "docs" / "data"

VILNIUS_TZ = ZoneInfo("Europe/Vilnius")

# Old "wide" layout headers (one row per station, header at row 8). ena.lt
# stopped serving this format / the direct-download URL in mid-2026.
OLD_HEADERS = [
    "Data",
    "Įmonė (Degalinių tinklas)",
    "Degalinės vieta (Savivaldybė)",
    "Degalinės vieta (Gyvenvietė, gatvė)",
    "95 benzinas",
    "Dyzelinas",
    "SND",
]

# New "long" layout headers (one row per station × fuel, header at row 7).
# Columns: Įmonė | Savivaldybė | Adresas | Degalų tipas | Kaina (EUR/l) |
# Pateikimo data. This is what the SharePoint share link now serves.
NEW_HEADERS = ["Įmonė", "Savivaldybė", "Adresas", "Degalų tipas", "Kaina (EUR/l)"]

# Maps the source's fuel-type label (new long format) to our JSON keys.
FUEL_TYPE_MAP = {
    "95 benzinas": "petrol95",
    "dyzelinas": "diesel",
    "snd": "lpg",
}

# Address-type / administrative filler tokens dropped when building the
# canonical station id. Both source formats include these (g.=gatvė,
# pr.=prospektas, k.=kaimas, sav.=savivaldybė, ...) but abbreviate them
# inconsistently, so dropping them makes the id robust to those differences.
_ID_FILLER_TOKENS = {
    "sav", "r", "m", "g", "pr", "k", "km", "mstl", "al", "pl",
    "plentas", "gatve", "gatvė", "prospektas", "kaimas", "kelias",
    "aikste", "aikštė", "a",
}
_POSTAL_RE = re.compile(r"\b\d{5}\b")
_TOKEN_RE = re.compile(r"[0-9a-ząčęėįšųūž]+")


def make_station_id(municipality: str, address: str) -> str:
    """Deterministic, format-independent ID from municipality + address.

    The source switched address formats in mid-2026: the old layout wrote
    `Gedimino g. 50 , Kėdainiai` while the new one writes
    `Kėdainiai, Gedimino g. 50, 57309` (city first, postal code appended).
    To keep one stable identity per physical station across that change we
    build the id from a sorted, de-duplicated set of address tokens with
    postal codes and address-type filler words stripped — so component order,
    the added postal code, and abbreviation differences all wash out.
    """
    raw = f"{municipality}|{address}".lower().replace("\xa0", " ")
    raw = _POSTAL_RE.sub(" ", raw)
    tokens = [t for t in _TOKEN_RE.findall(raw) if t not in _ID_FILLER_TOKENS]
    return "-".join(sorted(set(tokens)))


def parse_price(value) -> float | None:
    if isinstance(value, (int, float)):
        return round(float(value), 3)
    if isinstance(value, str):
        v = value.strip().replace(",", ".")
        if not v or v in ("-", "—", "–"):
            return None
        try:
            return round(float(v), 3)
        except ValueError:
            return None
    return None


def _find_header_row(ws, tokens: list[str], max_scan: int = 15) -> int | None:
    """Return the 1-based row index whose cells contain all `tokens`, or None."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        cells = {str(c).strip() for c in row if c is not None}
        if all(t in cells for t in tokens):
            return i
    return None


def _parse_long_format(ws, header_row: int) -> list[dict]:
    """Parse the new layout: one row per (station × fuel type).

    Rows look like: Įmonė | Savivaldybė | Adresas | Degalų tipas |
    Kaina (EUR/l) | Pateikimo data. We group consecutive rows for the same
    physical station (keyed by the canonical id) and fill its three fuel
    prices. Unknown fuel types and rows missing a company/address are skipped.
    """
    stations: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        company, municipality, address, fuel_type = row[0], row[1], row[2], row[3]
        if not company or not address:
            continue
        fuel = FUEL_TYPE_MAP.get(str(fuel_type).strip().lower())
        if fuel is None:
            continue
        sid = make_station_id(str(municipality), str(address))
        station = stations.get(sid)
        if station is None:
            station = {
                "id": sid,
                "company": company,
                "municipality": municipality,
                "address": address,
                "prices": {"petrol95": None, "diesel": None, "lpg": None},
            }
            stations[sid] = station
        station["prices"][fuel] = parse_price(row[4])
    return list(stations.values())


def _parse_wide_format(ws, header_row: int) -> list[dict]:
    """Parse the legacy layout: one row per station, three price columns."""
    headers = [cell.value for cell in list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]][:7]
    for i, expected in enumerate(OLD_HEADERS):
        if headers[i] != expected:
            print(f"WARNING: Column {i} header mismatch: expected '{expected}', got '{headers[i]}'")

    stations = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[1] is None:  # stop at summary/empty rows
            break
        municipality = row[2]
        address = row[3]
        stations.append({
            "id": make_station_id(str(municipality), str(address)),
            "company": row[1],
            "municipality": municipality,
            "address": address,
            "prices": {
                "petrol95": parse_price(row[4]),
                "diesel": parse_price(row[5]),
                "lpg": parse_price(row[6]),
            },
        })
    return stations


def parse_excel(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Degalų kainos"]
    try:
        # Detect layout by locating its header row. Prefer the new long format.
        new_row = _find_header_row(ws, ["Adresas", "Degalų tipas", "Kaina (EUR/l)"])
        if new_row is not None:
            return _parse_long_format(ws, new_row)
        old_row = _find_header_row(ws, ["Data", "95 benzinas", "Dyzelinas", "SND"])
        if old_row is not None:
            return _parse_wide_format(ws, old_row)
        raise ValueError("Could not locate a known header row in the Excel file")
    finally:
        wb.close()


def compute_price_changes(today: list[dict], yesterday_path: Path) -> dict:
    """Return {station_id: {petrol95: delta, diesel: delta, lpg: delta}}."""
    if not yesterday_path.exists():
        return {}

    with open(yesterday_path) as f:
        prev_data = json.load(f)

    prev_map = {s["id"]: s["prices"] for s in prev_data["stations"]}
    changes = {}

    for station in today:
        sid = station["id"]
        prev_prices = prev_map.get(sid)
        if not prev_prices:
            continue
        delta = {}
        for fuel in ("petrol95", "diesel", "lpg"):
            cur = station["prices"].get(fuel)
            prev = prev_prices.get(fuel)
            if cur is not None and prev is not None:
                delta[fuel] = round(cur - prev, 3)
            else:
                delta[fuel] = None
        changes[sid] = delta

    return changes


def compute_averages(stations: list[dict]) -> dict:
    sums = {"petrol95": 0, "diesel": 0, "lpg": 0}
    counts = {"petrol95": 0, "diesel": 0, "lpg": 0}
    for s in stations:
        for fuel in sums:
            p = s["prices"].get(fuel)
            if p is not None:
                sums[fuel] += p
                counts[fuel] += 1
    return {
        fuel: round(sums[fuel] / counts[fuel], 3) if counts[fuel] > 0 else None
        for fuel in sums
    }


def add_price_changes(stations: list[dict], changes: dict) -> list[dict]:
    """Add price change data to stations."""
    result = []
    for s in stations:
        entry = dict(s)
        entry["priceChange"] = changes.get(s["id"])
        result.append(entry)
    return result


def update_history_index():
    """Rebuild docs/data/history-index.json from data/history/ files."""
    dates = sorted([
        f.stem for f in HISTORY_DIR.glob("*.json")
    ])
    with open(DOCS_DATA / "history-index.json", "w") as f:
        json.dump(dates, f)


MINI_TREND_DAYS = 14


def attach_recent_prices(stations: list[dict], current_date: str) -> tuple[list[dict], list[str], dict]:
    """Embed last-N-days price history per station for popup sparklines.

    Reads the trailing MINI_TREND_DAYS history files (including `current_date`)
    and writes, per station, arrays aligned to a shared `recentDates` list.
    Missing days for a station become null. Returning enriched stations and
    the shared date list lets the frontend render a sparkline without any
    extra fetch.

    Also computes the Lithuania-wide average per fuel for each of those days
    so the popup sparkline can overlay a market-comparison line.
    """
    dates = sorted(f.stem for f in HISTORY_DIR.glob("*.json") if f.stem <= current_date)
    recent = dates[-MINI_TREND_DAYS:]
    if not recent:
        return stations, [], {fuel: [] for fuel in ("petrol95", "diesel", "lpg")}

    per_day: dict[str, list[dict]] = {}
    for d in recent:
        with open(HISTORY_DIR / f"{d}.json", encoding="utf-8") as f:
            hist = json.load(f)
        per_day[d] = hist["stations"]
    per_day_map = {d: {s["id"]: s["prices"] for s in per_day[d]} for d in recent}

    # Country-wide averages per fuel, aligned to `recent`.
    recent_averages = {fuel: [] for fuel in ("petrol95", "diesel", "lpg")}
    for d in recent:
        for fuel in recent_averages:
            vals = [s["prices"].get(fuel) for s in per_day[d] if s["prices"].get(fuel) is not None]
            recent_averages[fuel].append(round(sum(vals) / len(vals), 3) if vals else None)

    enriched = []
    for s in stations:
        sid = s["id"]
        recent_prices = {fuel: [] for fuel in ("petrol95", "diesel", "lpg")}
        for d in recent:
            day_prices = per_day_map[d].get(sid) or {}
            for fuel in recent_prices:
                recent_prices[fuel].append(day_prices.get(fuel))
        entry = dict(s)
        entry["recentPrices"] = recent_prices
        enriched.append(entry)
    return enriched, recent, recent_averages


BACKFILL_WINDOW_DAYS = 14


def _save_history(date_str: str, stations: list[dict]) -> None:
    HISTORY_DIR.mkdir(parents=True, exist_ok=True)
    path = HISTORY_DIR / f"{date_str}.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump({"date": date_str, "stations": stations}, f, ensure_ascii=False, indent=2)
    print(f"Saved {path}")


def _location_tokens(municipality: str, address: str) -> frozenset[str]:
    """Loose location signature for matching a station across id schemes.

    Unlike the canonical id this also drops single-letter tokens, so it spans
    differences the id doesn't (locality declension `Daugailiai`/`Daugailių`,
    an added city name, abbreviations). Used only to suppress ghost entries
    that are really the same physical station as a current one under a shifted
    id — never to merge stored data.
    """
    raw = f"{municipality}|{address}".lower().replace("\xa0", " ")
    raw = _POSTAL_RE.sub(" ", raw)
    return frozenset(t for t in _TOKEN_RE.findall(raw)
                     if t not in _ID_FILLER_TOKENS and len(t) > 1)


def build_recent_roster_ghosts(today_stations: list[dict], current_date: str,
                               window: int = MINI_TREND_DAYS) -> list[dict]:
    """Return 'ghost' entries for stations active recently but missing today.

    A station can drop out of the source for a day (or permanently — e.g. it
    closed or the operator stopped reporting). To keep the map roster stable
    we carry any station seen in the trailing `window` of history files but
    absent today as a priceless `missing` entry, so the frontend greys it out
    instead of letting it vanish. Metadata comes from its most recent
    appearance. Once a station is gone longer than the window it drops off.

    Ghosts that are really a current station under a shifted id (the source's
    mid-2026 address reformat changed some ids) are suppressed via a loose
    location match, so they don't appear as a duplicate beside their twin.
    """
    today_ids = {s["id"] for s in today_stations}
    today_sigs = [_location_tokens(s["municipality"], s["address"]) for s in today_stations]

    dates = sorted(f.stem for f in HISTORY_DIR.glob("*.json") if f.stem <= current_date)
    recent = dates[-window:]
    meta: dict[str, dict] = {}
    for d in recent:  # ascending → most recent appearance wins
        with open(HISTORY_DIR / f"{d}.json", encoding="utf-8") as f:
            for s in json.load(f)["stations"]:
                meta[s["id"]] = {
                    "company": s["company"],
                    "municipality": s["municipality"],
                    "address": s["address"],
                }

    def has_current_twin(sig: frozenset[str]) -> bool:
        # Same physical station if it shares the house number plus enough of
        # the street/locality with some current station.
        return any(len(sig & t) >= 3 and any(tok.isdigit() for tok in (sig & t))
                   for t in today_sigs)

    ghosts = []
    for sid, m in meta.items():
        if sid in today_ids:
            continue
        if has_current_twin(_location_tokens(m["municipality"], m["address"])):
            continue
        ghosts.append({**m, "id": sid,
                       "prices": {"petrol95": None, "diesel": None, "lpg": None},
                       "missing": True})
    return ghosts


def fetch_via_powerbi(target_date: datetime, date_str: str) -> bool:
    """Primary source: the LEA Power BI dataset embedded on ena.lt.

    It exposes the full daily history (verified identical to the Excel), so a
    single query both fetches the target day and self-heals any recent gap —
    every run backfills missing business days within the trailing window.
    Returns True if the target day's history file now exists.
    """
    from fetch_powerbi import stations_by_date

    start = (target_date - timedelta(days=BACKFILL_WINDOW_DAYS)).strftime("%Y-%m-%d")
    end = (target_date + timedelta(days=1)).strftime("%Y-%m-%d")
    by_date = stations_by_date(start, end, make_station_id, parse_price)
    backfilled = []
    for d, stations in sorted(by_date.items()):
        if not (HISTORY_DIR / f"{d}.json").exists():
            _save_history(d, list(stations.values()))
            backfilled.append(d)
    if backfilled:
        print(f"Power BI backfilled: {', '.join(backfilled)}")
    return (HISTORY_DIR / f"{date_str}.json").exists()


def fetch_via_sharepoint(date_str: str) -> list[dict]:
    """Fallback source: the SharePoint guest-share Excel (latest day only).

    ena.lt retired the direct .xlsx download URL in mid-2026; this is the
    backup if the Power BI dataset is unreachable. Exits cleanly if the
    target day isn't the latest published file.
    """
    from download_sharepoint import get_sharepoint_links, download_from_sharepoint_any

    sp_urls, sp_date = get_sharepoint_links(date_str)
    if sp_date != date_str:
        print(f"Latest available is {sp_date} but need {date_str} — not published yet, skipping")
        sys.exit(0)
    dl_dir = DATA_DIR / "downloads"
    dl_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = str(dl_dir / f"dk-{date_str}.xlsx")
    download_from_sharepoint_any(sp_urls, Path(xlsx_path))
    return parse_excel(xlsx_path)


def main():
    # Usage: fetch_prices.py [DATE] [--file PATH]
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("date", nargs="?", help="Target date YYYY-MM-DD")
    parser.add_argument("--file", help="Path to local Excel file (skip download)")
    args = parser.parse_args()

    if args.date:
        target_date = datetime.strptime(args.date, "%Y-%m-%d")
    else:
        target_date = datetime.now(VILNIUS_TZ)

    date_str = target_date.strftime("%Y-%m-%d")
    history_file = HISTORY_DIR / f"{date_str}.json"

    # Idempotent: skip if already fetched
    if history_file.exists():
        print(f"Data for {date_str} already exists, skipping fetch.")
        # Still regenerate stations.json in case geocache was updated
    else:
        if args.file:
            # Use provided local file
            print(f"Using local file: {args.file}")
            stations = parse_excel(args.file)
            print(f"Parsed {len(stations)} stations")
            _save_history(date_str, stations)
        else:
            # Primary: Power BI dataset (full history, self-backfilling).
            # Fall back to the SharePoint Excel if it's unreachable.
            got = False
            try:
                got = fetch_via_powerbi(target_date, date_str)
            except Exception as e:
                print(f"Power BI fetch failed ({e}); falling back to SharePoint")
            if not got:
                stations = fetch_via_sharepoint(date_str)
                print(f"Parsed {len(stations)} stations")
                _save_history(date_str, stations)

    # Load today's data
    with open(history_file, encoding="utf-8") as f:
        today_data = json.load(f)

    stations = today_data["stations"]

    # Find previous day's file for price changes
    prev_date = target_date - timedelta(days=1)
    # Search backwards up to 7 days for most recent previous data
    changes = {}
    for i in range(1, 8):
        check_date = target_date - timedelta(days=i)
        prev_file = HISTORY_DIR / f"{check_date.strftime('%Y-%m-%d')}.json"
        if prev_file.exists():
            changes = compute_price_changes(stations, prev_file)
            print(f"Computed price changes vs {check_date.strftime('%Y-%m-%d')}")
            break

    # Keep recently-active-but-missing stations on the map (greyed) instead of
    # letting them vanish. Averages stay over today's reporting stations only.
    ghosts = build_recent_roster_ghosts(stations, date_str)
    if ghosts:
        print(f"Roster: carrying {len(ghosts)} recently-active station(s) missing today as greyed")
    combined = stations + ghosts

    # Add price changes (ghosts have no current price → null change)
    enriched = add_price_changes(combined, changes)
    enriched, recent_dates, recent_averages = attach_recent_prices(enriched, date_str)
    averages = compute_averages(stations)

    frontend_data = {
        "date": date_str,
        "updatedAt": datetime.now(timezone.utc).isoformat(),
        "averages": averages,
        "recentDates": recent_dates,
        "recentAverages": recent_averages,
        "stations": enriched,
    }

    # Write latest.json and docs/data/stations.json
    with open(DATA_DIR / "latest.json", "w", encoding="utf-8") as f:
        json.dump(frontend_data, f, ensure_ascii=False, indent=2)

    DOCS_DATA.mkdir(parents=True, exist_ok=True)
    with open(DOCS_DATA / "stations.json", "w", encoding="utf-8") as f:
        json.dump(frontend_data, f, ensure_ascii=False, indent=2)

    print(f"Updated stations.json with {len(enriched)} stations")

    # Copy geocache to docs for frontend access
    geocache_path = DATA_DIR / "geocache.json"
    if geocache_path.exists():
        shutil.copy2(geocache_path, DOCS_DATA / "geocache.json")

    # Copy history files to docs for frontend access
    docs_history = DOCS_DATA / "history"
    docs_history.mkdir(parents=True, exist_ok=True)
    for src in HISTORY_DIR.glob("*.json"):
        shutil.copy2(src, docs_history / src.name)

    # Update history index
    update_history_index()
    print("Updated history-index.json")

    # Rebuild all derived trend artifacts (overall/network/region/station).
    # build_trends reads data/history/*.json from scratch each run, so it
    # also self-heals if any of the jsonl files drifted.
    import build_trends
    build_trends.main()


if __name__ == "__main__":
    main()
